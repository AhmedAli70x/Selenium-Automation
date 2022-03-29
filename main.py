# region Application Information
application_version = "1.0.0"
__author__ = "Ahme Ali"
__copyright__ = "Copyright © 2021"
__license__ = "GPL"
__version__ = application_version
__email__ = "ahmedali70x@gmail.com"
__status__ = "Production"

try:
    import os
    import zipfile
    from datetime import datetime
    from time import sleep
    from tkinter import *
    from tkinter import messagebox as mb
    from pprint import pprint

    import requests
    import wget
    from selenium import webdriver
    from selenium.webdriver.chrome.service import Service
    from selenium.webdriver.common.by import By
    from selenium.webdriver.common.keys import Keys
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.support.ui import WebDriverWait
    from openpyxl import workbook , load_workbook
    from openpyxl.utils import get_column_letter

    import config as cfg
except:
    def install(package):
        import subprocess
        subprocess.check_call(['pip3', "install", package])
    install('selenium')
    install('pywin32')
    install('requests')
    install('wget')
    install("openpyxl")
    import os
    import zipfile
    from datetime import datetime
    from time import sleep
    from tkinter import *
    from tkinter import messagebox as mb
    from pprint import pprint

    import requests
    import wget
    from selenium import webdriver
    from selenium.webdriver.chrome.service import Service
    from selenium.webdriver.common.by import By
    from selenium.webdriver.common.keys import Keys
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.support.ui import WebDriverWait
    from openpyxl import workbook , load_workbook
    from openpyxl.utils import get_column_letter

    import config as cfg


path_script = os.path.realpath(__file__)
if '/' in path_script:
    path_script = path_script[: path_script.rfind("/")]
    download_folder = path_script + f'/{cfg.DOWNLOAD_FOLDER}/'
    lib_folder = path_script + f'/{cfg.LIB_FOLDER}/'
    driver_path = lib_folder + 'chromedriver.exe'
else:
    path_script = path_script[: path_script.rfind("\\")]
    download_folder = path_script + fr'\{cfg.DOWNLOAD_FOLDER}' + '\\'
    lib_folder = path_script + fr'\{cfg.LIB_FOLDER}' + '\\'
    driver_path = lib_folder + 'chromedriver.exe'
user_profile = os.getenv('LOCALAPPDATA') + '\\'
root = Tk()
# endregion


def create_dir(path):
    isExist = os.path.exists(path)
    if not isExist:
        os.makedirs(path)


def download_chrome_driver():
    print(' [Status] Trying to update Chrome driver')
    print('*'*100)
    # get the latest chrome driver version number
    url = 'https://chromedriver.storage.googleapis.com/LATEST_RELEASE'
    response = requests.get(url)
    version_number = response.text

    # build the donwload url
    download_url = "https://chromedriver.storage.googleapis.com/" + version_number + "/chromedriver_win32.zip"

    # download the zip file using the url built above
    latest_driver_zip = wget.download(download_url, f'{lib_folder}chromedriver.zip')

    # extract the zip file
    with zipfile.ZipFile(latest_driver_zip, 'r') as zip_ref:
        zip_ref.extractall(lib_folder)  # you can specify the destination folder path here
    # delete the zip file downloaded above
    os.remove(latest_driver_zip)


def run_chrome():
    print(' [Status] Running Google Chrome')
    print('*'*100)

    options = webdriver.ChromeOptions()
    service = Service(driver_path)
    if cfg.HIDE_BROWSER:
        options.add_argument("--headless")
    options.add_argument("disable-gpu")
    options.add_argument("--start-maximized")
    options.add_argument('disable-infobars')
    options.add_argument("--disable-extensions")
    options.add_argument("--disable-notifications")
    options.add_experimental_option("prefs", {"profile.default_content_setting_values.notifications": 2})
    options.add_argument('log-level=3')
    if cfg.ENABLE_INCOGNITO:
        options.add_argument('--incognito')
    if cfg.BROWSER_NEW_PROFILE:
        options.add_argument(fr"user-data-dir={user_profile}Google\Chrome\User Data\{cfg.BROWSER_NEW_PROFILE}")
    options.add_experimental_option("excludeSwitches", ["disable-popup-blocking"])
    prefs = {'download.default_directory': download_folder}
    options.add_experimental_option('prefs', prefs)
    driver = webdriver.Chrome(options=options, service=service)

    if cfg.HIDE_BROWSER:
        driver.set_window_position(0, 0)
        driver.set_window_size(1920, 1080)

    return driver


def scrape_data(driver, url, year_week):

    driver.get(url)
    driver.refresh()
    delay = 4
    data = {}
    myElem = WebDriverWait(driver, delay).until(EC.presence_of_element_located((By.XPATH, '//*[@id="tm_table"]')))


    button_table = driver.find_elements(By.XPATH, '//*[@id="tm_table"]')
    if button_table:
        sleep(delay)
        button_table[0].click()
        myElem = WebDriverWait(driver, delay).until(EC.presence_of_element_located((By.XPATH, '//*[@id="tm_datatable"]')))
        sleep(delay/2)
        table_rows = driver.find_elements(By.XPATH, '//*[@id="tm_datatable"]/tbody/tr')
        list_data = []
        if table_rows:
            for row in table_rows:
                row_text = row.text.split(' ')
                code = row_text[0]
                res = row_text[-1]
                if 'disponible' in row_text or 'division' in row_text:
                    res = '#N/A'
                
                list_data.append({code: res})
            data[year_week] = list_data

            # print(data)

            return data


def check_column(ws, col_name):
    column_list =[]
    for cell in ws[1]:
        column_list.append(cell.value)
    
    return col_name in column_list

def insert_new_column(wb, ws, col_name, file):
    num_columns = ws.max_column
    new_column =  get_column_letter(num_columns + 1)
    new_year_week = new_column + str(1)
    ws[new_year_week].value = col_name
    wb.save(file)

def get_year_week_position(ws, col_name):
    index = 0 
    col_letter = False
    for col in ws[1]:
        index += 1
        if col.value == col_name:
            col_letter = get_column_letter(index)
    return col_letter 

def save_data(data, year_week, file):
    try:
        wb = load_workbook(file)
        ws  = wb.active
            
        check_year_week = check_column(ws, year_week)

        if not check_year_week:
            insert_new_column(wb, ws, year_week, file)

        col_letter = get_year_week_position(ws, year_week)
        
        if col_letter:
            for row in range(0, len(data[year_week])):
                for key, value in data[year_week][row].items():
                    # char = get_column_letter(row + 2)
                    sheet_code =  ws['A' + str(row + 2)]
                    if str(sheet_code.value) == key:
                        # print("True equal")
                        save_index = col_letter + str(row + 2)
                        # print(key, value)
                        value = value.replace(" ", "")
                        value = value.replace(" ", "")
                        if value.isnumeric():
                            value = int(value)
                        ws[save_index] = value
                        wb.save(file)
        return True

    except:
        print(f" [Status] Cannot save data for {year_week} while file is opened\n")
        return False

def main():
    create_dir(cfg.DOWNLOAD_FOLDER)
    create_dir(cfg.LIB_FOLDER)

    try:
        driver = run_chrome()
    except:
        download_chrome_driver()
        driver = run_chrome()

    for ind in cfg.INDICATORS:
        if ind.lower() in cfg.RESULT_ASTHMA_HOSPIT.lower():
            res_file = download_folder + cfg.RESULT_ASTHMA_HOSPIT
        elif ind.lower() in cfg.RESULT_ASTHMA_PASS.lower():
            res_file = download_folder + cfg.RESULT_ASTHMA_PASS

        for year, num_weeks in cfg.PERIOD.items():
            for week in range(1, num_weeks+1):
                year_week = "%04d-S%02d" % (year, week)
                url = f'https://geodes.santepubliquefrance.fr/#c=indicator&f=0&i=sursaud_sau.prop_asthme_{ind}_sau&s={year_week}&t=a01&view=map2'
                try:
                    print(f" [Status] Scraping url {url}")
                    data = scrape_data(driver, url, year_week)
                    if data:
                        print(f" [Status] Saving data")
                        # print(res_file)
                        save = save_data(data, year_week, res_file)
                        if save:
                            print(f" [Status] Data for {year_week} is saved successfully\n")
                except:
                    wb = load_workbook(res_file)
                    ws  = wb.active
                    check_year_week = check_column(ws, year_week)
                    if not check_year_week:
                        insert_new_column(wb, ws, year_week, res_file)

                    print(f" [Status] Error found in URL: {url}\n")
                    with open("Failed.txt", "a+") as text_file:
                        text_file.write(f'{url}\n')
                    text_file.close()

                if cfg.TEST:
                    driver.quit()
                    exit()

    print(" [Status] Scraping is complete...")
    driver.quit()


if __name__ == "__main__":
    main()
